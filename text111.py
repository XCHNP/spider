import json
import os
import random
import re
import time
import openpyxl
import requests
from lxml import etree
from functools import reduce

# 从文档中读出已经获取的url，爬取所有url的具体内容并存储

user_agent_list = [
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) Gecko/20100101 Firefox/61.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.62 Safari/537.36",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
    "Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10.5; en-US; rv:1.9.2.15) Gecko/20110303 Firefox/3.6.15",
]
headers = {'User-Agent': random.choice(user_agent_list)}


# 修改url最后一位达到翻页效果
def replace_lastchar(former_str, replacechar):
    return former_str[:-6] + replacechar + '.html'


def requests_2(basicurl, num):  # 对二层页面的所有页面解析，并获得具体页的url
    with open('urls.txt', 'a+', encoding='utf-8') as fp:
        for i in range(1, num + 1):
            # 修改url最后一位访问所有页面
            url = replace_lastchar(basicurl, str(i))
            geturl3 = requests.get(url=url, headers=headers).text
            tree3 = etree.HTML(geturl3)
            tr_list = tree3.xpath('//*[contains(@id,"normalthread")]')  # 返回值为列表

            # 找到所有具体页面的url写入文件
            for tr in tr_list:
                fp.write('https://www.5ilaobing.com/' + tr.xpath('./tr/th/a[3]/@href')[0])
                # print("abc")
                fp.write('\r\n')


def getpagenum_2(url):  # 对二层页面内容解析，获得翻页页数
    geturl2 = requests.get(url=url, headers=headers).text
    tree2 = etree.HTML(geturl2)
    pagenums = tree2.xpath('//*[@id="fd_page_bottom"]/div/label/span/@title')

    # 提取出总页码
    # 1.将score_list列表转换为以“，”为分隔符的字符串
    pagenums_string = ",".join(pagenums)
    # 2.使用re模块下的方法——findall对name_score_list_string进行正则匹配，提取出所有的数值型字符串
    score_list = re.findall(r"\d+", pagenums_string)
    # 3.将得到的score_list中的元素转换为int型
    pagenums_int = list(map(int, score_list))
    if len(pagenums_int) == 0:
        pagenums_int.append(0)

    requests_2(url, pagenums_int[0])


def pc1():
    # 解析第一个网页
    url = 'https://www.5ilaobing.com/'
    geturl1 = requests.get(url=url, headers=headers).text
    tree = etree.HTML(geturl1)
    tr_list = tree.xpath('//*[@id="category_5"]//tr')
    # 获得全部具体地址的url，得到url2s
    url2s = []

    for tr in tr_list:
        if len(tr.xpath('./td[2]/h2/a/@href')) == 1:
            url2s.append('https://www.5ilaobing.com/' + tr.xpath('./td[2]/h2/a/@href')[0])

    # 循环访问二层页面
    for url in url2s:
        getpagenum_2(url)

    print("爬取成功！")


detail_lists = []


# 具体页面解析
def final_requests(detail_url, i):
    time.sleep(0.5)
    page_text = requests.get(url=detail_url, headers=headers).text

    # 数据解析
    tree = etree.HTML(page_text)
    content = tree.xpath('//*[@class="t_f"]/text()')
    # print(content)
    print('get!' + str(i))
    return content


# --------文本分析1----------------数据清洗------------------

def get_index(l, n):  # 查找冒号前空格在文字前最后一次出现的下标
    for i in range(0, n):
        if l[i] != ' ':
            return i
        else:
            return False


def refresh(tempstr):  # 去除数据内部乱码
    detail_str = tempstr.replace('\n', '')
    detail_str = detail_str.replace('\r', '')
    detail_str = detail_str.replace(' ', '')
    detail_str = detail_str.replace('</P>', '')
    detail_str = detail_str.replace(' >', '')
    detail_str = detail_str.replace(u'\xa0', '')  # 去除&nbsp表示的空格
    return detail_str


def use_ch(use_str):  # 去除字符串中非中文
    n1 = len(use_str)
    for i in range(0, n1):
        pattern = re.compile(r'[^\u4e00-\u9fa5]')
        changed_str = re.sub(pattern, '', use_str)
        return changed_str


def getNumofCommonSubstr(str1, str2):
    lstr1 = len(str1)
    lstr2 = len(str2)
    record = [[0 for i in range(lstr2 + 1)] for j in range(lstr1 + 1)]
    # 开辟列表空间 为什么要多一位呢?主要是不多一位的话,会存在边界问题
    # 多了一位以后就不存在超界问题
    maxNum = 0  # 最长匹配长度
    p = 0  # 匹配的起始位

    for i in range(lstr1):
        for j in range(lstr2):
            if str1[i] == str2[j]:
                # 相同则累加
                record[i + 1][j + 1] = record[i][j] + 1
                if record[i + 1][j + 1] > maxNum:
                    # 获取最大匹配长度
                    maxNum = record[i + 1][j + 1]
                    # 记录最大匹配长度的终止位置
                    p = i + 1
    return str1[p - maxNum:p]


# ------------文本分析2-------规律文本转列表------------------


dic_list = []
allname_index = []
allcontent_index = []
def deal_text(details):  # 文本分析主程序
    # 去除列表中的空元素
    while '' in details:
        details.remove('')

    name_index = []  # 将冒号前的所有内容存进本列表
    content_index = []  # 将冒号后的所有内容存进本列表
    n1 = len(details)

    for i in range(0, n1): # 循环到列表每个元素
        n = len(details[i]) #循环每个元素的字符串
        # 处理冒号前元素
        if details[i].find('：') != -1:  # 有冒号元素
            detail_str = refresh(details[i])
            head_index = get_index(detail_str, n)
            end_index = detail_str.find('：')
            if len(detail_str[head_index:end_index]) < 15:  # 去除大段描述性语句
                final_str = use_ch(detail_str[head_index:end_index])
                if len(allname_index) > 0 and final_str!=None:
                    index = 0
                    for allname in allname_index:
                        for j in range(0, len(allname)):
                            if len(getNumofCommonSubstr(final_str, allname[j])) > 3:  # 重合度高，需要取子集
                                name_index.append(getNumofCommonSubstr(final_str, allname[j]))
                                allname[j] = getNumofCommonSubstr(final_str, allname[j])  # 将之前出现过的替换为子串
                                index = 1
                                break
                        if index==1:
                            break
                    if index == 0:
                        name_index.append(final_str)
                elif final_str==None:
                    continue
                else:
                    name_index.append(final_str)
            # 处理冒号后元素
            if detail_str[detail_str.find('：') + 1:] == '':
                content_index.append('null')
            else:
                content_index.append(detail_str[detail_str.find('：') + 1:])

    allname_index.append(name_index)
    allcontent_index.append(content_index)



# 每个字典中，找出有集合中的键对应的值放入对应的列表中，否则放一个空值进去(要保证都有对应关系)


# 应该是可以用哈希，写一大半想起来了................................靠..............无语.....................

occ_str = 'null'


def deal_key(dic, final_list, keys):  # 处理键作为集合用于后续

    for key in keys:
        index = keys.index(key)
        if dic.get(key, 0) != 0:  # 确认本字典中有该键对应的值
            final_list[index].append(dic[key])
        else:
            final_list[index].append(occ_str)
    return final_list


def into_dic():
    # 将处理好的字典写入文件
    with open('./detail_data.txt', 'a+', encoding='utf-8') as f:
        for dic in dic_list:
            json_str = json.dumps(dic, ensure_ascii=False)  # dumps
            f.write(json_str)
            f.write('\n')


# -------------文本处理3--------------其余文字处理------保留列表中无冒号的部分，jieba分词，保留关键信息

other_lists = []

def deal_other(details):  # 保留列表中无冒号的部分
    other_list = []
    n1 = len(details)
    for i in range(0, n1):
        n = len(details[i])
        if details[i].find('：') == -1 and len(details[i]) > 50:  # 无冒号元素
            detail_str = refresh(details[i])
            other_list.append(detail_str)

    other_list = ''.join(other_list)
    if len(other_list) == 0:
        other_list = []
        for j in range(0, n1):
            n = len(details[j])
            if len(details[j]) > 50:  # 无冒号元素
                other_list.append(details[j])
        other_list = ''.join(other_list)
        if len(other_list) == 0:
            other_lists.append('null')
        else:
            other_lists.append(other_list)
    else:
        other_lists.append(other_list)


def into_excel(file_path):
    keys.append('其他信息')
    final_list.append(other_lists)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '老兵信息'
    for i in range(len(keys)):
        ws.cell(1, i + 1).value = keys[i]
    for r in range(len(final_list)):
        for c in range(len(final_list[r])):
            ws.cell(c + 2, r + 1).value = final_list[r][c]
            # excel中的行和列是从1开始计数的，所以需要+1
    wb.save(file_path)  # 注意，写入后一定要保存
    print("成功写入文件: " + file_path + " !")
    return 1


if __name__ == '__main__':
    # 从文件中读取已爬好的url，并逐条请求获得详情页
    urls = []
    if os.access("urls.txt", os.F_OK) == 0:
        urls = pc1()
        print('已获取全部url！')

    print('文件已存在!')
    with open('./urls.txt', "r", encoding="utf-8") as f:
        line = f.readline()  # 读取第一行
        while line:
            line = line.strip('\n')  # 去除每行末尾的换行符
            urls.append(line)  # 列表增加
            line = f.readline()  # 读取下一行

    print('已获取全部url！')

    # 爬取+初步处理

    for i in range(7000, 7500):  # !!!!在这里可以修改运行一次获取的个数，如果一次跑不完就多分几次跑(不过要记得修改excel名)
        if i % 10 == 0:
            time.sleep(10)
        if urls[i].startswith('h'):
            contents = final_requests(urls[i], i)
            deal_text(contents)
            deal_other(contents)

    print('全部爬取成功！')

    # print(dic_list)
    into_dic()


    temp_list = reduce(lambda x, y: x + y, allname_index) # 二维列表转一维
    keys = list(set(temp_list))
    # print(type(keys))
    # print(set(temp_list))
    # print(keys)

    # 把姓名作为第一列
    if keys[0].startswith('姓') == 0:
        for i in range(1, len(keys)):
            if keys[i].startswith('姓') != 0:
                temp = keys[0]
                keys[0] = keys[i]
                keys[i] = temp

    # print(keys)
    for i in range(0, len(allname_index)):
        dic_list.append(dict(zip(allname_index[i], allcontent_index[i])))

    length = len(keys)
    final_list = [[] for i in range(length)]  # 存储最终所有的值，元素为存储相同值的列表  初始化为二维空列表
    for dic in dic_list:
        deal_key(dic, final_list, keys)
    # print(final_list)  # 输出规律部分转化的结果

    while '' in other_lists:
        other_lists.remove('')
    into_excel('./detail15.xls')
    # print(final_list)
    # print(dic_list)
    # print(len(other_lists))
    print(keys)  # 输出列名，不对了自己填吧....
    print('程序结束！')
